from __future__ import annotations

import csv
import os
import re
import subprocess
import tempfile
from io import BytesIO, StringIO
from zipfile import BadZipFile
import fitz  # PyMuPDF

from docx import Document as DocxDocument
from openpyxl import load_workbook
from pdf2image import convert_from_bytes
from PIL import Image
from pypdf import PdfReader
import pytesseract
import xlrd

from app.core.config import settings


class ParseResult(dict):
    text: str
    used_ocr: bool
    parser: str


SUPPORTED_EXTENSIONS = {
    '.pdf', '.png', '.jpg', '.jpeg', '.webp', '.tif', '.tiff', '.bmp',
    '.heic', '.heif',
    '.doc', '.docx', '.xls', '.xlsx', '.csv', '.txt'
}

SUPPORTED_MIME_TYPES = {
    'application/pdf',
    'image/png', 'image/jpeg', 'image/jpg', 'image/webp', 'image/tiff', 'image/bmp',
    'image/heic', 'image/heif',
    'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
    'application/msword',
    'application/vnd.ms-excel',
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    'text/csv', 'application/csv',
    'text/plain',
    'application/octet-stream',
}


def get_extension(filename: str | None) -> str:
    return os.path.splitext(filename or '')[1].lower()


def is_supported_upload(filename: str | None, mime_type: str | None) -> bool:
    extension = get_extension(filename)
    mime = (mime_type or '').lower().strip()
    if extension in SUPPORTED_EXTENSIONS:
        return True
    return mime in SUPPORTED_MIME_TYPES


def _extract_pdf_text(content: bytes) -> str:
    text = ""
    try:
        with fitz.open(stream=content, filetype="pdf") as doc:
            for page in doc:
                text += page.get_text()
    except Exception:
        pass
    return text.strip()


def _extract_docx_text(content: bytes) -> str:
    doc = DocxDocument(BytesIO(content))
    return '\n'.join(p.text for p in doc.paragraphs).strip()


def _extract_doc_text(content: bytes) -> str:
    with tempfile.NamedTemporaryFile(suffix='.doc', delete=True) as tmp:
        tmp.write(content)
        tmp.flush()
        completed = subprocess.run(
            ['antiword', tmp.name],
            capture_output=True,
            text=True,
            timeout=30,
            check=False,
        )
        text = (completed.stdout or '').strip()
        if text:
            return text
        raise ValueError((completed.stderr or 'Could not extract legacy Word document text').strip())


def _extract_image_text(content: bytes) -> str:
    image = Image.open(BytesIO(content))
    return pytesseract.image_to_string(image).strip()


def _ocr_pdf(content: bytes) -> str:
    images = convert_from_bytes(content, dpi=200)
    return '\n'.join(pytesseract.image_to_string(image).strip() for image in images).strip()


def _extract_csv_text(content: bytes) -> str:
    decoded = content.decode('utf-8', errors='ignore')
    rows = list(csv.reader(StringIO(decoded)))
    return '\n'.join(' | '.join(cell.strip() for cell in row if cell is not None) for row in rows).strip()


def _extract_xlsx_text(content: bytes) -> str:
    wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
    lines: list[str] = []
    for sheet in wb.worksheets:
        lines.append(f'Sheet: {sheet.title}')
        for row in sheet.iter_rows(values_only=True):
            values = [str(cell).strip() for cell in row if cell not in (None, '')]
            if values:
                lines.append(' | '.join(values))
    return '\n'.join(lines).strip()


def _extract_xls_text(content: bytes) -> str:
    book = xlrd.open_workbook(file_contents=content)
    lines: list[str] = []
    for idx in range(book.nsheets):
        sheet = book.sheet_by_index(idx)
        lines.append(f'Sheet: {sheet.name}')
        for rx in range(sheet.nrows):
            values = [str(value).strip() for value in sheet.row_values(rx) if str(value).strip()]
            if values:
                lines.append(' | '.join(values))
    return '\n'.join(lines).strip()


def _extract_plain_text(content: bytes) -> str:
    text = content.decode('utf-8', errors='ignore').strip()
    text = re.sub(r'\x00+', ' ', text)
    return text.strip()


def parse_file(content: bytes, mime_type: str, filename: str | None = None) -> ParseResult:
    text = ''
    used_ocr = False
    parser = 'plain_text'
    mime = (mime_type or '').lower().strip()
    extension = get_extension(filename)

    try:
        if mime == 'application/pdf' or extension == '.pdf':
            parser = 'pdf_text'
            text = _extract_pdf_text(content)

            if (not text or not text.strip()) and settings.OCR_ENABLED:
                parser = 'pdf_ocr'
                text = _ocr_pdf(content)
                used_ocr = bool(text and text.strip())

        elif mime in {'application/vnd.openxmlformats-officedocument.wordprocessingml.document'} or extension == '.docx':
            parser = 'docx'
            text = _extract_docx_text(content)

        elif mime == 'application/msword' or extension == '.doc':
            parser = 'doc'
            text = _extract_doc_text(content)

        elif mime in {'application/vnd.ms-excel'} or extension == '.xls':
            parser = 'xls'
            text = _extract_xls_text(content)

        elif mime in {'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'} or extension == '.xlsx':
            parser = 'xlsx'
            text = _extract_xlsx_text(content)

        elif mime in {'text/csv', 'application/csv'} or extension == '.csv':
            parser = 'csv'
            text = _extract_csv_text(content)

        elif mime.startswith('image/') or extension in {'.png', '.jpg', '.jpeg', '.webp', '.tif', '.tiff', '.bmp'}:
            parser = 'image_ocr'
            if settings.OCR_ENABLED:
                text = _extract_image_text(content)
                used_ocr = bool(text and text.strip())

        else:
            parser = 'plain_text'
            text = _extract_plain_text(content)

    except (BadZipFile, ValueError, xlrd.XLRDError, subprocess.TimeoutExpired):
        parser = 'plain_text_fallback'
        text = _extract_plain_text(content)
    except Exception:
        parser = 'plain_text_fallback'
        text = _extract_plain_text(content)

    cleaned_text = (text or '').strip()

    return {
        'text': cleaned_text if cleaned_text else 'No extractable text found.',
        'used_ocr': used_ocr,
        'parser': parser,
    }